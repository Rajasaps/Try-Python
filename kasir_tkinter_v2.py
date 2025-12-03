import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from datetime import datetime
import json
import os
from PIL import Image, ImageTk
from db_config import Database, save_transaction, get_all_transactions, get_transaction_stats

class KedaiHaunaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Kedai Hauna - Aplikasi Kasir")
        self.root.geometry("1400x800")
        self.root.configure(bg="#0a0a0a")
        
        # Colors (sama dengan web)
        self.colors = {
            'bg_dark': '#0a0a0a',
            'bg_card': '#1a1a1a',
            'bg_sidebar': '#000000',
            'text_white': '#ffffff',
            'text_gray': '#999999',
            'accent': '#ff4444',
            'border': '#333333'
        }
        
        # Data menu
        self.menu_items = [
            {"id": 1, "name": "Bakso Malang", "price": 23000, "image": "static/images/Bakso malang.jpg"},
            {"id": 2, "name": "Seblak Special", "price": 16000, "image": "static/images/Seblak special.jpg"},
            {"id": 3, "name": "Mie Ayam", "price": 18000, "image": "static/images/Mie Ayam.jpg"},
            {"id": 4, "name": "Siomay", "price": 11000, "image": "static/images/Siomay.jpg"},
            {"id": 5, "name": "Tea", "price": 6000, "image": "static/images/Tea.jpg", "has_variant": True, "variants": ["Dingin", "Hangat"]},
            {"id": 6, "name": "Ayam Crispy", "price": 17000, "image": "static/images/Ayam Crispy.jpg"},
            {"id": 7, "name": "Nasi", "price": 5000, "image": "static/images/Nasi.jpg"},
        ]
        
        self.cart = []
        self.images = {}
        
        # Initialize database connection
        self.db = Database()
        if not self.db.connect():
            messagebox.showerror("Database Error", 
                               "Tidak dapat terhubung ke database!\n\n"
                               "Pastikan:\n"
                               "1. XAMPP sudah running\n"
                               "2. MySQL service aktif\n"
                               "3. Database 'kedai_hauna' sudah dibuat\n\n"
                               "Aplikasi akan menggunakan mode offline (JSON).")
            self.db = None
        
        self.transactions = self.load_transactions()
        
        self.load_images()
        self.setup_ui()
        self.update_cart_display()  # Initialize cart display
        
        # Notification queue and history
        self.notification_queue = []
        self.notification_history = []
    
    def bind_mousewheel(self, canvas):
        """Bind mouse wheel to canvas for scrolling"""
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def on_enter(event):
            canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        def on_leave(event):
            canvas.unbind_all("<MouseWheel>")
        
        canvas.bind("<Enter>", on_enter)
        canvas.bind("<Leave>", on_leave)
    
    def show_notification(self, message, duration=3000, type="success"):
        """Tampilkan notifikasi toast di pojok kanan atas"""
        # Save to history
        self.notification_history.insert(0, {
            'message': message,
            'type': type,
            'time': datetime.now().strftime('%H:%M:%S')
        })
        
        # Keep only last 50 notifications
        if len(self.notification_history) > 50:
            self.notification_history = self.notification_history[:50]
        
        # Create notification window
        notif = tk.Toplevel(self.root)
        notif.overrideredirect(True)  # Remove window decorations
        notif.attributes('-topmost', True)  # Always on top
        
        # Set background color based on type
        bg_color = "#4CAF50" if type == "success" else "#ff4444" if type == "error" else "#2196F3"
        
        # Create frame
        frame = tk.Frame(notif, bg=bg_color, padx=20, pady=15)
        frame.pack()
        
        # Icon based on type
        icon = "✓" if type == "success" else "✗" if type == "error" else "ℹ"
        
        # Icon label
        tk.Label(frame, text=icon, font=("Arial", 16, "bold"), 
                bg=bg_color, fg="white").pack(side=tk.LEFT, padx=(0, 10))
        
        # Message label
        tk.Label(frame, text=message, font=("Segoe UI", 11), 
                bg=bg_color, fg="white", wraplength=300).pack(side=tk.LEFT)
        
        # Position at top right
        notif.update_idletasks()
        width = notif.winfo_width()
        height = notif.winfo_height()
        
        # Calculate position (top right with offset)
        x = self.root.winfo_x() + self.root.winfo_width() - width - 20
        y = self.root.winfo_y() + 20 + (len(self.notification_queue) * (height + 10))
        
        notif.geometry(f"+{x}+{y}")
        
        # Add to queue
        self.notification_queue.append(notif)
        
        # Fade in animation
        notif.attributes('-alpha', 0.0)
        self.fade_in(notif, 0.0)
        
        # Auto close after duration
        def close_notification():
            self.fade_out(notif, 1.0)
            if notif in self.notification_queue:
                self.notification_queue.remove(notif)
            # Reposition remaining notifications
            self.reposition_notifications()
        
        notif.after(duration, close_notification)
    
    def fade_in(self, window, alpha):
        """Fade in animation"""
        if alpha < 1.0:
            alpha += 0.1
            window.attributes('-alpha', alpha)
            window.after(30, lambda: self.fade_in(window, alpha))
    
    def fade_out(self, window, alpha):
        """Fade out animation"""
        if alpha > 0.0:
            alpha -= 0.1
            window.attributes('-alpha', alpha)
            window.after(30, lambda: self.fade_out(window, alpha))
        else:
            window.destroy()
    
    def reposition_notifications(self):
        """Reposition notifications after one is closed"""
        for idx, notif in enumerate(self.notification_queue):
            notif.update_idletasks()
            width = notif.winfo_width()
            height = notif.winfo_height()
            x = self.root.winfo_x() + self.root.winfo_width() - width - 20
            y = self.root.winfo_y() + 20 + (idx * (height + 10))
            notif.geometry(f"+{x}+{y}")
    
    def show_notification_history(self):
        """Tampilkan history notifikasi"""
        # Create dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Riwayat Notifikasi")
        dialog.geometry("500x600")
        dialog.configure(bg=self.colors['bg_card'])
        dialog.transient(self.root)
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # Header
        header = tk.Frame(dialog, bg=self.colors['bg_card'])
        header.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(header, text="🔔 Riwayat Notifikasi", font=("Segoe UI", 14, "bold"),
                bg=self.colors['bg_card'], fg="white").pack(side=tk.LEFT)
        
        tk.Button(header, text="×", font=("Arial", 18), bg=self.colors['bg_card'],
                 fg="white", relief=tk.FLAT, command=dialog.destroy).pack(side=tk.RIGHT)
        
        # Clear all button
        if self.notification_history:
            tk.Button(header, text="Hapus Semua", font=("Segoe UI", 10), 
                     bg=self.colors['accent'], fg="white", relief=tk.FLAT,
                     command=lambda: self.clear_notification_history(dialog),
                     padx=10, pady=5).pack(side=tk.RIGHT, padx=10)
        
        # Scrollable list
        canvas = tk.Canvas(dialog, bg=self.colors['bg_card'], highlightthickness=0)
        scrollbar = tk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        list_frame = tk.Frame(canvas, bg=self.colors['bg_card'])
        
        list_frame.bind("<Configure>", 
                       lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=list_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=10)
        
        # Enable mouse wheel scrolling
        self.bind_mousewheel(canvas)
        
        # Display notifications
        if not self.notification_history:
            tk.Label(list_frame, text="Belum ada notifikasi", font=("Segoe UI", 11),
                    bg=self.colors['bg_card'], fg=self.colors['text_gray']).pack(pady=50)
        else:
            for notif in self.notification_history:
                self.create_notification_item(list_frame, notif)
    
    def create_notification_item(self, parent, notif):
        """Create notification item in history"""
        item_frame = tk.Frame(parent, bg=self.colors['bg_dark'])
        item_frame.pack(fill=tk.X, pady=5, padx=5)
        
        # Icon based on type
        icon = "✓" if notif['type'] == "success" else "✗" if notif['type'] == "error" else "ℹ"
        icon_color = "#4CAF50" if notif['type'] == "success" else "#ff4444" if notif['type'] == "error" else "#2196F3"
        
        # Left: Icon
        tk.Label(item_frame, text=icon, font=("Arial", 14, "bold"),
                bg=self.colors['bg_dark'], fg=icon_color, width=3).pack(side=tk.LEFT, padx=10, pady=10)
        
        # Middle: Message
        msg_frame = tk.Frame(item_frame, bg=self.colors['bg_dark'])
        msg_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=10)
        
        tk.Label(msg_frame, text=notif['message'], font=("Segoe UI", 10),
                bg=self.colors['bg_dark'], fg="white", anchor="w", wraplength=350).pack(fill=tk.X)
        tk.Label(msg_frame, text=notif['time'], font=("Segoe UI", 9),
                bg=self.colors['bg_dark'], fg=self.colors['text_gray'], anchor="w").pack(fill=tk.X)
    
    def clear_notification_history(self, dialog):
        """Clear all notification history"""
        if messagebox.askyesno("Konfirmasi", "Hapus semua riwayat notifikasi?"):
            self.notification_history = []
            dialog.destroy()
            self.show_notification("Riwayat notifikasi dihapus", duration=2000, type="info")
    
    def load_images(self):
        """Load dan resize gambar menu"""
        for item in self.menu_items:
            try:
                if os.path.exists(item['image']):
                    img = Image.open(item['image'])
                    # Resize untuk card yang lebih kecil
                    img = img.resize((200, 150), Image.Resampling.LANCZOS)
                    self.images[item['id']] = ImageTk.PhotoImage(img)
            except Exception as e:
                print(f"Error loading {item['image']}: {e}")
    
    def setup_ui(self):
        # Container utama
        self.main_container = tk.Frame(self.root, bg=self.colors['bg_dark'])
        self.main_container.pack(fill=tk.BOTH, expand=True)
        
        # Sidebar kiri
        self.setup_sidebar(self.main_container)
        
        # Content area tengah
        self.content_frame = tk.Frame(self.main_container, bg=self.colors['bg_dark'])
        self.content_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Cart sidebar akan dibuat di show_kasir_page
        self.cart_sidebar = None
        
        # Show kasir page
        self.show_kasir_page()

    def setup_sidebar(self, parent):
        sidebar = tk.Frame(parent, bg=self.colors['bg_sidebar'], width=200)
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)
        
        # Logo
        logo_frame = tk.Frame(sidebar, bg=self.colors['bg_sidebar'])
        logo_frame.pack(pady=25, padx=15)
        
        # Logo icon dengan background merah rounded
        logo_bg = tk.Frame(logo_frame, bg=self.colors['accent'], width=50, height=50)
        logo_bg.pack()
        logo_bg.pack_propagate(False)
        
        logo_icon = tk.Label(logo_bg, text="🍜", font=("Arial", 24), 
                            bg=self.colors['accent'], fg="white")
        logo_icon.pack(expand=True)
        
        logo_text = tk.Label(logo_frame, text="Kedai\nHauna", font=("Segoe UI", 13, "bold"), 
                            bg=self.colors['bg_sidebar'], fg="white", justify=tk.LEFT)
        logo_text.pack(pady=8)
        
        # Menu buttons
        menu_frame = tk.Frame(sidebar, bg=self.colors['bg_sidebar'])
        menu_frame.pack(fill=tk.BOTH, expand=True, padx=10)
        
        self.btn_kasir = tk.Button(menu_frame, text="💰 Kasir", font=("Segoe UI", 11), 
                                   bg="white", fg=self.colors['accent'], relief=tk.FLAT,
                                   command=self.show_kasir_page, anchor="w", padx=18, pady=13,
                                   cursor="hand2", borderwidth=0)
        self.btn_kasir.pack(fill=tk.X, pady=4)
        
        self.btn_pemasukan = tk.Button(menu_frame, text="📊 Pemasukan", font=("Segoe UI", 11),
                                       bg=self.colors['bg_card'], fg=self.colors['text_gray'], 
                                       relief=tk.FLAT, command=self.show_pemasukan_page, 
                                       anchor="w", padx=18, pady=13, cursor="hand2", borderwidth=0)
        self.btn_pemasukan.pack(fill=tk.X, pady=4)
        
        # Logout button
        tk.Button(sidebar, text="🚪 Logout", font=("Segoe UI", 11), 
                 bg=self.colors['bg_card'], fg=self.colors['text_gray'], relief=tk.FLAT,
                 command=self.root.quit, anchor="w", padx=18, pady=13, cursor="hand2",
                 borderwidth=0).pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=20)

    def setup_cart_sidebar(self, parent):
        self.cart_sidebar = tk.Frame(parent, bg=self.colors['bg_sidebar'], width=300)
        self.cart_sidebar.pack(side=tk.RIGHT, fill=tk.Y)
        self.cart_sidebar.pack_propagate(False)
        cart_sidebar = self.cart_sidebar  # Keep local variable for compatibility
        
        # Logo card - Compact for small screens
        logo_frame = tk.Frame(cart_sidebar, bg=self.colors['bg_card'])
        logo_frame.pack(fill=tk.X, padx=15, pady=10)
        
        # Logo Kedai Hauna - Smaller
        logo_bg = tk.Frame(logo_frame, bg=self.colors['accent'], width=45, height=45)
        logo_bg.pack(pady=8)
        logo_bg.pack_propagate(False)
        
        tk.Label(logo_bg, text="🍜", font=("Arial", 24), 
                bg=self.colors['accent'], fg="white").pack(expand=True)
        
        tk.Label(logo_frame, text="Kedai Hauna", font=("Segoe UI", 12, "bold"),
                bg=self.colors['bg_card'], fg="white").pack()
        tk.Label(logo_frame, text="📍 Purbalingga, Indonesia", font=("Segoe UI", 8),
                bg=self.colors['bg_card'], fg=self.colors['text_gray']).pack(pady=3)
        
        # Customer name input - Compact
        customer_frame = tk.Frame(logo_frame, bg=self.colors['bg_card'])
        customer_frame.pack(fill=tk.X, padx=10, pady=8)
        
        tk.Label(customer_frame, text="Nama Customer:", font=("Segoe UI", 9),
                bg=self.colors['bg_card'], fg=self.colors['text_gray'], anchor="w").pack(fill=tk.X, pady=(0, 3))
        
        self.customer_name_var = tk.StringVar()
        customer_entry = tk.Entry(customer_frame, textvariable=self.customer_name_var,
                                  font=("Segoe UI", 10), bg=self.colors['bg_dark'],
                                  fg="white", relief=tk.FLAT, insertbackground="white")
        customer_entry.pack(fill=tk.X, ipady=6, padx=2)
        
        # Cart section
        cart_section = tk.Frame(cart_sidebar, bg=self.colors['bg_card'])
        cart_section.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)
        
        # Cart header - Compact
        cart_header = tk.Frame(cart_section, bg=self.colors['bg_card'])
        cart_header.pack(fill=tk.X, padx=15, pady=5)
        
        tk.Label(cart_header, text="Keranjang Belanja", font=("Arial", 11, "bold"),
                bg=self.colors['bg_card'], fg="white").pack(side=tk.LEFT)
        tk.Button(cart_header, text="🗑️", font=("Arial", 11), bg=self.colors['bg_card'],
                 fg="white", relief=tk.FLAT, command=self.clear_cart).pack(side=tk.RIGHT)
        
        # Cart items scrollable - Increased height to show more items
        cart_scroll_container = tk.Frame(cart_section, bg=self.colors['bg_card'], height=280)
        cart_scroll_container.pack(fill=tk.X, padx=15)
        cart_scroll_container.pack_propagate(False)
        
        canvas = tk.Canvas(cart_scroll_container, bg=self.colors['bg_card'], highlightthickness=0)
        scrollbar = tk.Scrollbar(cart_scroll_container, orient="vertical", command=canvas.yview)
        self.cart_items_frame = tk.Frame(canvas, bg=self.colors['bg_card'])
        
        self.cart_items_frame.bind("<Configure>", 
                                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.cart_items_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Enable mouse wheel scrolling
        self.bind_mousewheel(canvas)
        
        # Summary - Very compact for small screens
        summary_frame = tk.Frame(cart_section, bg=self.colors['bg_card'])
        summary_frame.pack(fill=tk.X, padx=15, pady=5)
        
        self.subtotal_label = tk.Label(summary_frame, text="Subtotal: Rp 0", 
                                       font=("Arial", 8), bg=self.colors['bg_card'], 
                                       fg="white", anchor="w")
        self.subtotal_label.pack(fill=tk.X, pady=1)
        
        self.tax_label = tk.Label(summary_frame, text="Pajak (10%): Rp 0",
                                  font=("Arial", 8), bg=self.colors['bg_card'],
                                  fg="white", anchor="w")
        self.tax_label.pack(fill=tk.X, pady=1)
        
        tk.Frame(summary_frame, bg=self.colors['border'], height=1).pack(fill=tk.X, pady=3)
        
        self.total_label = tk.Label(summary_frame, text="Total: Rp 0",
                                    font=("Arial", 11, "bold"), bg=self.colors['bg_card'],
                                    fg="white", anchor="w")
        self.total_label.pack(fill=tk.X, pady=2)
        
        tk.Button(summary_frame, text="Bayar Sekarang", font=("Arial", 10, "bold"),
                 bg=self.colors['accent'], fg="white", relief=tk.FLAT,
                 command=self.checkout, pady=8).pack(fill=tk.X, pady=5)

    def show_kasir_page(self):
        # Update button styles
        self.btn_kasir.config(bg="white", fg=self.colors['accent'])
        self.btn_pemasukan.config(bg=self.colors['bg_card'], fg=self.colors['text_gray'])
        
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        # Show cart sidebar (only in kasir page)
        if self.cart_sidebar:
            self.cart_sidebar.destroy()
        self.setup_cart_sidebar(self.main_container)
        
        # Header dengan search
        header = tk.Frame(self.content_frame, bg=self.colors['bg_dark'])
        header.pack(fill=tk.X, padx=20, pady=15)
        
        search_frame = tk.Frame(header, bg=self.colors['bg_card'])
        search_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        tk.Label(search_frame, text="🔍", font=("Arial", 12), 
                bg=self.colors['bg_card'], fg="white").pack(side=tk.LEFT, padx=10)
        
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, font=("Arial", 11), bg=self.colors['bg_card'],
                               fg="white", relief=tk.FLAT, insertbackground="white",
                               textvariable=self.search_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=12, padx=5)
        
        # Placeholder text
        placeholder_text = "Cari menu..."
        search_entry.insert(0, placeholder_text)
        search_entry.config(fg=self.colors['text_gray'])
        
        def on_focus_in(event):
            if search_entry.get() == placeholder_text:
                search_entry.delete(0, tk.END)
                search_entry.config(fg="white")
        
        def on_focus_out(event):
            if search_entry.get() == "":
                search_entry.insert(0, placeholder_text)
                search_entry.config(fg=self.colors['text_gray'])
        
        search_entry.bind("<FocusIn>", on_focus_in)
        search_entry.bind("<FocusOut>", on_focus_out)
        
        # Icons
        icon_frame = tk.Frame(header, bg=self.colors['bg_dark'])
        icon_frame.pack(side=tk.RIGHT, padx=10)
        
        # Notification icon (clickable)
        notif_btn = tk.Label(icon_frame, text="🔔", font=("Arial", 16), bg=self.colors['bg_dark'],
                            fg="white", cursor="hand2")
        notif_btn.pack(side=tk.LEFT, padx=8)
        notif_btn.bind("<Button-1>", lambda e: self.show_notification_history())
        
        tk.Label(icon_frame, text="👤", font=("Arial", 16), bg=self.colors['bg_dark'],
                fg="white").pack(side=tk.LEFT, padx=8)
        
        # Menu grid dengan scroll
        canvas = tk.Canvas(self.content_frame, bg=self.colors['bg_dark'], highlightthickness=0)
        scrollbar = tk.Scrollbar(self.content_frame, orient="vertical", command=canvas.yview)
        menu_container = tk.Frame(canvas, bg=self.colors['bg_dark'])
        
        # Bind untuk update scroll region
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        menu_container.bind("<Configure>", on_frame_configure)
        
        # Create window dengan width yang akan di-update
        canvas_window = canvas.create_window((0, 0), window=menu_container, anchor="nw")
        
        # Bind untuk update window width saat canvas resize
        def on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)
        
        canvas.bind("<Configure>", on_canvas_configure)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Enable mouse wheel scrolling
        self.bind_mousewheel(canvas)
        
        # Configure grid columns untuk 3 kolom
        menu_container.grid_columnconfigure(0, weight=1, uniform="col")
        menu_container.grid_columnconfigure(1, weight=1, uniform="col")
        menu_container.grid_columnconfigure(2, weight=1, uniform="col")
        
        # Store menu cards untuk search
        self.menu_cards = {}
        
        # Create menu cards (3 kolom)
        for idx, item in enumerate(self.menu_items):
            row = idx // 3
            col = idx % 3
            card_widget = self.create_menu_card(menu_container, item, row, col)
            self.menu_cards[item['id']] = {'widget': card_widget, 'item': item}
        
        # No results message (initially hidden)
        self.no_results_frame = tk.Frame(menu_container, bg=self.colors['bg_dark'])
        self.no_results_frame.grid(row=100, column=0, columnspan=3, pady=50)
        self.no_results_frame.grid_remove()
        
        tk.Label(self.no_results_frame, text="🔍", font=("Arial", 48),
                bg=self.colors['bg_dark'], fg=self.colors['text_gray']).pack(pady=10)
        tk.Label(self.no_results_frame, text="Menu tidak ditemukan", font=("Segoe UI", 16, "bold"),
                bg=self.colors['bg_dark'], fg="white").pack(pady=5)
        tk.Label(self.no_results_frame, text="Coba kata kunci lain untuk mencari menu", font=("Segoe UI", 11),
                bg=self.colors['bg_dark'], fg=self.colors['text_gray']).pack(pady=5)
        
        # Bind search function
        def search_menu(*args):
            search_text = self.search_var.get().lower()
            if search_text == "cari menu...":
                search_text = ""
            
            visible_count = 0
            for item_id, card_data in self.menu_cards.items():
                item_name = card_data['item']['name'].lower()
                if search_text in item_name or search_text == "":
                    card_data['widget'].grid()
                    visible_count += 1
                else:
                    card_data['widget'].grid_remove()
            
            # Show/hide no results message
            if visible_count == 0 and search_text != "":
                self.no_results_frame.grid()
            else:
                self.no_results_frame.grid_remove()
        
        self.search_var.trace('w', search_menu)

    def create_menu_card(self, parent, item, row, col):
        # Outer frame untuk rounded effect
        card_outer = tk.Frame(parent, bg=self.colors['bg_dark'], highlightthickness=0)
        card_outer.grid(row=row, column=col, padx=12, pady=12, sticky="nsew")
        
        # Store grid info untuk show/hide
        card_outer.grid_info_stored = {'row': row, 'column': col, 'padx': 12, 'pady': 12, 'sticky': 'nsew'}
        
        card = tk.Frame(card_outer, bg=self.colors['bg_card'], highlightthickness=0)
        card.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Gambar dengan rounded corners
        if item['id'] in self.images:
            img_label = tk.Label(card, image=self.images[item['id']], bg=self.colors['bg_card'])
            img_label.pack(fill=tk.X, pady=0)
        else:
            # Placeholder jika gambar tidak ada
            placeholder = tk.Label(card, text="🍜", font=("Arial", 60),
                                  bg="#333333", fg="white", height=3)
            placeholder.pack(fill=tk.X)
        
        # Info frame dengan relative positioning untuk button
        info_container = tk.Frame(card, bg=self.colors['bg_card'])
        info_container.pack(fill=tk.BOTH, expand=True)
        
        # Text info
        info_frame = tk.Frame(info_container, bg=self.colors['bg_card'])
        info_frame.pack(fill=tk.X, padx=16, pady=(12, 8))
        
        tk.Label(info_frame, text=item['name'], font=("Segoe UI", 13, "bold"),
                bg=self.colors['bg_card'], fg="white", anchor="w").pack(fill=tk.X)
        tk.Label(info_frame, text=f"Rp {item['price']:,}", font=("Segoe UI", 11),
                bg=self.colors['bg_card'], fg=self.colors['text_gray'], 
                anchor="w").pack(fill=tk.X, pady=(4, 0))
        
        # Add button positioned at bottom right
        btn_container = tk.Frame(info_container, bg=self.colors['bg_card'])
        btn_container.pack(fill=tk.X, padx=16, pady=(0, 16))
        
        if item.get("has_variant"):
            add_btn = tk.Button(btn_container, text="+", font=("Arial", 18, "bold"),
                               bg="white", fg="black", relief=tk.FLAT, 
                               width=2, height=1, cursor="hand2",
                               command=lambda i=item: self.show_variant_dialog(i))
        else:
            add_btn = tk.Button(btn_container, text="+", font=("Arial", 18, "bold"),
                               bg="white", fg="black", relief=tk.FLAT,
                               width=2, height=1, cursor="hand2",
                               command=lambda i=item: self.add_to_cart(i))
        add_btn.pack(side=tk.RIGHT)
        
        # Hover effect
        def on_enter(e):
            card.config(bg="#222222")
            info_container.config(bg="#222222")
            info_frame.config(bg="#222222")
            btn_container.config(bg="#222222")
            for child in info_frame.winfo_children():
                child.config(bg="#222222")
        
        def on_leave(e):
            card.config(bg=self.colors['bg_card'])
            info_container.config(bg=self.colors['bg_card'])
            info_frame.config(bg=self.colors['bg_card'])
            btn_container.config(bg=self.colors['bg_card'])
            for child in info_frame.winfo_children():
                child.config(bg=self.colors['bg_card'])
        
        card.bind("<Enter>", on_enter)
        card.bind("<Leave>", on_leave)
        if item['id'] in self.images:
            img_label.bind("<Enter>", on_enter)
            img_label.bind("<Leave>", on_leave)
        
        return card_outer

    def show_variant_dialog(self, item):
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Pilih Varian")
        dialog.geometry("350x250")
        dialog.configure(bg=self.colors['bg_card'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # Header
        header = tk.Frame(dialog, bg=self.colors['bg_card'])
        header.pack(fill=tk.X, padx=20, pady=20)
        
        tk.Label(header, text="Pilih Metode Pembayaran", font=("Arial", 13, "bold"),
                bg=self.colors['bg_card'], fg="white").pack(side=tk.LEFT)
        tk.Button(header, text="×", font=("Arial", 16), bg=self.colors['bg_card'],
                 fg="white", relief=tk.FLAT, command=dialog.destroy).pack(side=tk.RIGHT)
        
        # Variant buttons
        for variant in item.get("variants", []):
            btn = tk.Button(dialog, text=variant, font=("Arial", 12),
                           bg=self.colors['bg_dark'], fg="white", relief=tk.FLAT,
                           command=lambda v=variant: self.add_variant_and_close(item, v, dialog),
                           pady=15)
            btn.pack(fill=tk.X, padx=20, pady=5)
            
            # Hover effect
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg=self.colors['accent']))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg=self.colors['bg_dark']))
    
    def add_variant_and_close(self, item, variant, dialog):
        item_copy = item.copy()
        item_copy["variant"] = variant
        item_copy["name"] = f"{item['name']} ({variant})"
        self.add_to_cart(item_copy)
        dialog.destroy()
    
    def add_to_cart(self, item):
        item_key = f"{item['id']}_{item.get('variant', '')}"
        
        for cart_item in self.cart:
            if cart_item.get("item_key") == item_key:
                cart_item["quantity"] += 1
                self.update_cart_display()
                # Show notification
                self.show_notification(f"✓ {item['name']} ditambahkan ke keranjang", duration=2000)
                return
        
        self.cart.append({
            "id": item["id"],
            "item_key": item_key,
            "name": item["name"],
            "price": item["price"],
            "quantity": 1,
            "variant": item.get("variant", "")
        })
        self.update_cart_display()
        # Show notification
        self.show_notification(f"✓ {item['name']} ditambahkan ke keranjang", duration=2000)

    def update_cart_display(self):
        for widget in self.cart_items_frame.winfo_children():
            widget.destroy()
        
        if not self.cart:
            tk.Label(self.cart_items_frame, text="Keranjang kosong", font=("Arial", 10),
                    bg=self.colors['bg_card'], fg="#666666").pack(pady=30)
        else:
            for item in self.cart:
                self.create_cart_item(item)
        
        # Update summary
        subtotal = sum(item["price"] * item["quantity"] for item in self.cart)
        tax = subtotal * 0.1
        total = subtotal + tax
        
        self.subtotal_label.config(text=f"Subtotal: Rp {subtotal:,.0f}")
        self.tax_label.config(text=f"Pajak (10%): Rp {tax:,.0f}")
        self.total_label.config(text=f"Total: Rp {total:,.0f}")
    
    def create_cart_item(self, item):
        item_frame = tk.Frame(self.cart_items_frame, bg=self.colors['bg_card'])
        item_frame.pack(fill=tk.X, pady=3)
        
        # Right: Remove button (pack first for proper alignment)
        remove_btn = tk.Button(item_frame, text="×", font=("Arial", 14, "bold"), 
                               bg=self.colors['accent'], fg="white", relief=tk.FLAT, 
                               width=2, command=lambda: self.remove_from_cart(item))
        remove_btn.pack(side=tk.RIGHT, padx=3)
        
        # Left: Info
        info_frame = tk.Frame(item_frame, bg=self.colors['bg_card'])
        info_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        tk.Label(info_frame, text=item["name"], font=("Arial", 9, "bold"),
                bg=self.colors['bg_card'], fg="white", anchor="w").pack(fill=tk.X)
        tk.Label(info_frame, text=f"Rp {item['price']:,}", font=("Arial", 8),
                bg=self.colors['bg_card'], fg=self.colors['text_gray'], 
                anchor="w").pack(fill=tk.X)
        
        # Quantity controls - Compact
        qty_frame = tk.Frame(info_frame, bg=self.colors['bg_card'])
        qty_frame.pack(fill=tk.X, pady=3)
        
        tk.Button(qty_frame, text="-", font=("Arial", 8), bg="#333333", fg="white",
                 relief=tk.FLAT, width=2, command=lambda: self.update_quantity(item, -1)).pack(side=tk.LEFT)
        tk.Label(qty_frame, text=str(item["quantity"]), font=("Arial", 9, "bold"),
                bg=self.colors['bg_card'], fg="white", width=3).pack(side=tk.LEFT, padx=3)
        tk.Button(qty_frame, text="+", font=("Arial", 8), bg="#333333", fg="white",
                 relief=tk.FLAT, width=2, command=lambda: self.update_quantity(item, 1)).pack(side=tk.LEFT)
    
    def update_quantity(self, item, change):
        item["quantity"] += change
        if item["quantity"] <= 0:
            self.cart.remove(item)
        self.update_cart_display()
    
    def remove_from_cart(self, item):
        self.cart.remove(item)
        self.update_cart_display()
    
    def clear_cart(self):
        if messagebox.askyesno("Konfirmasi", "Hapus semua item dari keranjang?"):
            self.cart = []
            self.update_cart_display()

    def checkout(self):
        if not self.cart:
            messagebox.showwarning("Peringatan", "Keranjang masih kosong!")
            return
        
        subtotal = sum(item["price"] * item["quantity"] for item in self.cart)
        tax = subtotal * 0.1
        total = subtotal + tax
        
        # Payment dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Pilih Metode Pembayaran")
        dialog.geometry("450x600")
        dialog.configure(bg=self.colors['bg_card'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # Header
        header = tk.Frame(dialog, bg=self.colors['bg_card'])
        header.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(header, text="Pilih Metode Pembayaran", font=("Arial", 14, "bold"),
                bg=self.colors['bg_card'], fg="white").pack(side=tk.LEFT)
        tk.Button(header, text="×", font=("Arial", 18), bg=self.colors['bg_card'],
                 fg="white", relief=tk.FLAT, command=dialog.destroy).pack(side=tk.RIGHT)
        
        # Summary
        summary = tk.Frame(dialog, bg=self.colors['bg_dark'])
        summary.pack(fill=tk.X, padx=20, pady=10)
        
        tk.Label(summary, text=f"Subtotal: Rp {subtotal:,.0f}", font=("Arial", 10),
                bg=self.colors['bg_dark'], fg="white", anchor="w").pack(fill=tk.X, padx=15, pady=3)
        tk.Label(summary, text=f"Pajak (10%): Rp {tax:,.0f}", font=("Arial", 10),
                bg=self.colors['bg_dark'], fg="white", anchor="w").pack(fill=tk.X, padx=15, pady=3)
        tk.Label(summary, text=f"Total Bayar: Rp {total:,.0f}", font=("Arial", 14, "bold"),
                bg=self.colors['bg_dark'], fg="white", anchor="w").pack(fill=tk.X, padx=15, pady=8)
        
        # Payment methods
        payment_var = tk.StringVar(value="cash")
        
        methods_frame = tk.Frame(dialog, bg=self.colors['bg_card'])
        methods_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        methods = [
            ("💵 Tunai", "cash"),
            ("💳 Kartu Debit", "debit"),
            ("💳 Kartu Kredit", "credit"),
            ("📱 E-Wallet", "ewallet"),
            ("📲 QRIS", "qris"),
            ("🏦 Transfer Bank", "transfer")
        ]
        
        for text, value in methods:
            rb = tk.Radiobutton(methods_frame, text=text, variable=payment_var, value=value,
                               font=("Arial", 11), bg=self.colors['bg_card'], fg="white",
                               selectcolor="#333333", activebackground=self.colors['bg_card'],
                               activeforeground="white", anchor="w", padx=10, pady=8)
            rb.pack(fill=tk.X, pady=3)
        
        # Confirm button
        tk.Button(dialog, text="Konfirmasi Pembayaran", font=("Arial", 12, "bold"),
                 bg=self.colors['accent'], fg="white", relief=tk.FLAT, pady=12,
                 command=lambda: self.process_payment(payment_var.get(), total, dialog)).pack(
                 fill=tk.X, padx=20, pady=20)

    def process_payment(self, method, total, dialog):
        if method == "cash":
            cash_dialog = tk.Toplevel(dialog)
            cash_dialog.title("Pembayaran Tunai")
            cash_dialog.geometry("400x300")
            cash_dialog.configure(bg=self.colors['bg_card'])
            cash_dialog.transient(dialog)
            cash_dialog.grab_set()
            
            tk.Label(cash_dialog, text="Pembayaran Tunai", font=("Arial", 14, "bold"),
                    bg=self.colors['bg_card'], fg="white").pack(pady=20)
            
            tk.Label(cash_dialog, text=f"Total: Rp {total:,.0f}", font=("Arial", 12),
                    bg=self.colors['bg_card'], fg="white").pack(pady=10)
            
            tk.Label(cash_dialog, text="Masukkan jumlah uang:", font=("Arial", 11),
                    bg=self.colors['bg_card'], fg="white").pack(pady=5)
            
            cash_entry = tk.Entry(cash_dialog, font=("Arial", 14), bg=self.colors['bg_dark'],
                                 fg="white", justify="center", insertbackground="white")
            cash_entry.pack(pady=10, padx=40, fill=tk.X)
            cash_entry.focus()
            
            change_label = tk.Label(cash_dialog, text="Kembalian: Rp 0", font=("Arial", 12, "bold"),
                                   bg=self.colors['bg_card'], fg="#4CAF50")
            change_label.pack(pady=10)
            
            def calculate_change(*args):
                try:
                    # Remove spaces and convert to int
                    cash_str = cash_entry.get().strip().replace(" ", "").replace(",", "")
                    if cash_str:
                        cash = int(cash_str)
                        change = cash - total
                        if change >= 0:
                            change_label.config(text=f"Kembalian: Rp {change:,.0f}", fg="#4CAF50")
                        else:
                            change_label.config(text=f"Kurang: Rp {abs(change):,.0f}", fg=self.colors['accent'])
                    else:
                        change_label.config(text="Kembalian: Rp 0", fg="#4CAF50")
                except:
                    change_label.config(text="Kembalian: Rp 0", fg="#4CAF50")
            
            cash_entry.bind("<KeyRelease>", calculate_change)
            
            def confirm_cash():
                try:
                    # Remove spaces and convert to int
                    cash_str = cash_entry.get().strip().replace(" ", "").replace(",", "")
                    if not cash_str:
                        messagebox.showerror("Error", "Masukkan jumlah uang!")
                        return
                    
                    cash = int(cash_str)
                    if cash < total:
                        messagebox.showerror("Error", f"Uang tidak cukup!\nKurang: Rp {total - cash:,.0f}")
                        return
                    cash_dialog.destroy()
                    self.finalize_payment(method, total, cash, dialog)
                except ValueError:
                    messagebox.showerror("Error", "Masukkan jumlah uang yang valid!\nHanya angka yang diperbolehkan.")
                except Exception as e:
                    messagebox.showerror("Error", f"Terjadi kesalahan: {str(e)}")
            
            tk.Button(cash_dialog, text="Konfirmasi", font=("Arial", 12, "bold"),
                     bg=self.colors['accent'], fg="white", relief=tk.FLAT, pady=10,
                     command=confirm_cash).pack(fill=tk.X, padx=40, pady=20)
        else:
            self.finalize_payment(method, total, 0, dialog)
    
    def finalize_payment(self, method, total, cash_amount, dialog):
        # Get customer name
        customer_name = self.customer_name_var.get() if hasattr(self, 'customer_name_var') and self.customer_name_var.get() else "Umum"
        
        transaction_id = datetime.now().strftime('%Y%m%d%H%M%S')
        
        transaction = {
            "id": transaction_id,
            "items": self.cart.copy(),
            "total": total,
            "payment_method": method,
            "customer_name": customer_name,
            "date": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Save to database
        if self.db:
            try:
                if save_transaction(self.db, transaction_id, customer_name, method, self.cart, total):
                    self.show_notification("✓ Transaksi tersimpan ke database", duration=2000, type="success")
                else:
                    self.show_notification("⚠ Gagal simpan ke database, tersimpan di JSON", duration=3000, type="error")
            except Exception as e:
                print(f"Error saving to database: {e}")
                self.show_notification("⚠ Error database, tersimpan di JSON", duration=3000, type="error")
        
        self.transactions.append(transaction)
        self.save_transactions()  # Backup to JSON
        
        dialog.destroy()
        
        payment_names = {
            'cash': 'Tunai', 'debit': 'Kartu Debit', 'credit': 'Kartu Kredit',
            'ewallet': 'E-Wallet', 'qris': 'QRIS', 'transfer': 'Transfer Bank'
        }
        
        receipt = f"✓ Pembayaran Berhasil!\n\n"
        receipt += f"ID Transaksi: {transaction['id']}\n"
        receipt += f"Metode: {payment_names[method]}\n"
        receipt += f"Total: Rp {total:,.0f}\n"
        
        if method == "cash" and cash_amount > 0:
            change = cash_amount - total
            receipt += f"\nUang Tunai: Rp {cash_amount:,.0f}\n"
            receipt += f"Kembalian: Rp {change:,.0f}\n"
        
        receipt += "\nTerima kasih atas pembelian Anda!"
        
        # Show messagebox
        messagebox.showinfo("Pembayaran Berhasil", receipt)
        
        # Show notification
        notif_message = f"Pesanan berhasil dibayar! ID: {transaction['id']}"
        self.show_notification(notif_message, duration=4000, type="success")
        
        # Ask if want to print receipt
        if messagebox.askyesno("Cetak Nota", "Apakah Anda ingin mencetak nota?"):
            self.print_receipt(transaction, cash_amount if method == "cash" else 0)
        
        self.cart = []
        self.update_cart_display()
    
    def print_receipt(self, transaction, cash_amount=0):
        """Generate dan tampilkan nota untuk dicetak"""
        # Create receipt window
        receipt_window = tk.Toplevel(self.root)
        receipt_window.title(f"Nota - {transaction['id']}")
        receipt_window.geometry("400x700")
        receipt_window.configure(bg="white")
        
        # Center window
        receipt_window.update_idletasks()
        x = (receipt_window.winfo_screenwidth() // 2) - (receipt_window.winfo_width() // 2)
        y = (receipt_window.winfo_screenheight() // 2) - (receipt_window.winfo_height() // 2)
        receipt_window.geometry(f"+{x}+{y}")
        
        # Receipt content frame
        content_frame = tk.Frame(receipt_window, bg="white", padx=30, pady=20)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header
        tk.Label(content_frame, text="KEDAI HAUNA", font=("Courier New", 16, "bold"),
                bg="white", fg="black").pack()
        tk.Label(content_frame, text="Purbalingga, Indonesia", font=("Courier New", 9),
                bg="white", fg="black").pack()
        tk.Label(content_frame, text="=" * 40, font=("Courier New", 10),
                bg="white", fg="black").pack(pady=5)
        
        # Transaction info
        info_frame = tk.Frame(content_frame, bg="white")
        info_frame.pack(fill=tk.X, pady=10)
        
        tk.Label(info_frame, text=f"No. Transaksi: {transaction['id']}", 
                font=("Courier New", 9), bg="white", fg="black", anchor="w").pack(fill=tk.X)
        tk.Label(info_frame, text=f"Tanggal: {transaction['date']}", 
                font=("Courier New", 9), bg="white", fg="black", anchor="w").pack(fill=tk.X)
        
        # Customer name
        customer_name = self.customer_name_var.get() if hasattr(self, 'customer_name_var') and self.customer_name_var.get() else "Umum"
        tk.Label(info_frame, text=f"Customer: {customer_name}", 
                font=("Courier New", 9), bg="white", fg="black", anchor="w").pack(fill=tk.X)
        
        tk.Label(content_frame, text="-" * 40, font=("Courier New", 10),
                bg="white", fg="black").pack(pady=5)
        
        # Items
        items_frame = tk.Frame(content_frame, bg="white")
        items_frame.pack(fill=tk.X)
        
        for item in transaction['items']:
            item_line = tk.Frame(items_frame, bg="white")
            item_line.pack(fill=tk.X, pady=2)
            
            tk.Label(item_line, text=item['name'], font=("Courier New", 9),
                    bg="white", fg="black", anchor="w").pack(side=tk.LEFT)
            
            price_text = f"{item['quantity']}x {item['price']:,} = {item['quantity'] * item['price']:,}"
            tk.Label(item_line, text=price_text, font=("Courier New", 9),
                    bg="white", fg="black", anchor="e").pack(side=tk.RIGHT)
        
        tk.Label(content_frame, text="-" * 40, font=("Courier New", 10),
                bg="white", fg="black").pack(pady=5)
        
        # Summary
        summary_frame = tk.Frame(content_frame, bg="white")
        summary_frame.pack(fill=tk.X)
        
        subtotal = transaction['total'] / 1.1
        tax = transaction['total'] - subtotal
        
        tk.Label(summary_frame, text=f"Subtotal: Rp {subtotal:,.0f}", 
                font=("Courier New", 9), bg="white", fg="black", anchor="w").pack(fill=tk.X)
        tk.Label(summary_frame, text=f"Pajak (10%): Rp {tax:,.0f}", 
                font=("Courier New", 9), bg="white", fg="black", anchor="w").pack(fill=tk.X)
        
        tk.Label(content_frame, text="=" * 40, font=("Courier New", 10),
                bg="white", fg="black").pack(pady=5)
        
        tk.Label(content_frame, text=f"TOTAL: Rp {transaction['total']:,.0f}", 
                font=("Courier New", 12, "bold"), bg="white", fg="black").pack()
        
        # Payment info
        payment_names = {
            'cash': 'Tunai', 'debit': 'Kartu Debit', 'credit': 'Kartu Kredit',
            'ewallet': 'E-Wallet', 'qris': 'QRIS', 'transfer': 'Transfer Bank'
        }
        
        payment_frame = tk.Frame(content_frame, bg="white")
        payment_frame.pack(fill=tk.X, pady=10)
        
        tk.Label(payment_frame, text=f"Metode: {payment_names.get(transaction['payment_method'], 'Unknown')}", 
                font=("Courier New", 9), bg="white", fg="black", anchor="w").pack(fill=tk.X)
        
        if cash_amount > 0:
            tk.Label(payment_frame, text=f"Bayar: Rp {cash_amount:,.0f}", 
                    font=("Courier New", 9), bg="white", fg="black", anchor="w").pack(fill=tk.X)
            tk.Label(payment_frame, text=f"Kembali: Rp {cash_amount - transaction['total']:,.0f}", 
                    font=("Courier New", 9), bg="white", fg="black", anchor="w").pack(fill=tk.X)
        
        tk.Label(content_frame, text="=" * 40, font=("Courier New", 10),
                bg="white", fg="black").pack(pady=5)
        
        # Footer
        tk.Label(content_frame, text="Terima Kasih", font=("Courier New", 10, "bold"),
                bg="white", fg="black").pack(pady=5)
        tk.Label(content_frame, text="Selamat Datang Kembali", font=("Courier New", 9),
                bg="white", fg="black").pack()
        
        # Buttons
        button_frame = tk.Frame(receipt_window, bg="white", pady=10)
        button_frame.pack(fill=tk.X)
        
        tk.Button(button_frame, text="💾 Simpan PDF", font=("Segoe UI", 10, "bold"),
                 bg="#4CAF50", fg="white", relief=tk.FLAT, padx=20, pady=10,
                 command=lambda: self.save_receipt_pdf(transaction, cash_amount, receipt_window)).pack(side=tk.LEFT, padx=10)
        
        tk.Button(button_frame, text="🖨️ Print", font=("Segoe UI", 10, "bold"),
                 bg="#2196F3", fg="white", relief=tk.FLAT, padx=20, pady=10,
                 command=lambda: self.print_receipt_direct(transaction, cash_amount)).pack(side=tk.LEFT, padx=10)
        
        tk.Button(button_frame, text="Tutup", font=("Segoe UI", 10),
                 bg="#999", fg="white", relief=tk.FLAT, padx=20, pady=10,
                 command=receipt_window.destroy).pack(side=tk.RIGHT, padx=10)
    
    def save_receipt_pdf(self, transaction, cash_amount, window):
        """Simpan nota sebagai PDF file"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.units import mm
            from reportlab.pdfgen import canvas
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            filename = f"nota_{transaction['id']}.pdf"
            
            # Create PDF
            c = canvas.Canvas(filename, pagesize=A4)
            width, height = A4
            
            # Set up fonts
            try:
                # Try to use system fonts for better Unicode support
                c.setFont("Helvetica-Bold", 16)
            except:
                c.setFont("Helvetica-Bold", 16)
            
            # Header
            y = height - 50
            c.setFont("Helvetica-Bold", 20)
            c.drawCentredString(width/2, y, "KEDAI HAUNA")
            
            y -= 20
            c.setFont("Helvetica", 10)
            c.drawCentredString(width/2, y, "Purbalingga, Indonesia")
            
            # Line separator
            y -= 15
            c.line(50, y, width-50, y)
            
            # Transaction info
            y -= 30
            c.setFont("Helvetica-Bold", 12)
            c.drawString(50, y, "NOTA PEMBAYARAN")
            
            y -= 25
            c.setFont("Helvetica", 10)
            c.drawString(50, y, f"No. Transaksi:")
            c.drawString(200, y, transaction['id'])
            
            y -= 20
            c.drawString(50, y, f"Tanggal:")
            c.drawString(200, y, transaction['date'])
            
            y -= 20
            customer_name = transaction.get('customer_name', 'Umum')
            c.drawString(50, y, f"Customer:")
            c.drawString(200, y, customer_name.title())
            
            payment_names = {
                'cash': 'Tunai', 'debit': 'Kartu Debit', 'credit': 'Kartu Kredit',
                'ewallet': 'E-Wallet', 'qris': 'QRIS', 'transfer': 'Transfer Bank'
            }
            
            y -= 20
            c.drawString(50, y, f"Metode Bayar:")
            c.drawString(200, y, payment_names.get(transaction['payment_method'], 'Unknown'))
            
            # Line separator
            y -= 15
            c.line(50, y, width-50, y)
            
            # Items header
            y -= 25
            c.setFont("Helvetica-Bold", 10)
            c.drawString(50, y, "Item")
            c.drawString(300, y, "Qty")
            c.drawString(360, y, "Harga")
            c.drawString(470, y, "Subtotal")
            
            y -= 5
            c.line(50, y, width-50, y)
            
            # Items
            y -= 20
            c.setFont("Helvetica", 9)
            for item in transaction['items']:
                if y < 100:  # Check if we need a new page
                    c.showPage()
                    y = height - 50
                    c.setFont("Helvetica", 9)
                
                # Item name (wrap if too long)
                item_name = item['name']
                if len(item_name) > 30:
                    item_name = item_name[:27] + "..."
                c.drawString(50, y, item_name)
                
                # Quantity
                c.drawString(300, y, str(item['quantity']))
                
                # Price
                c.drawString(360, y, f"Rp {item['price']:,}")
                
                # Subtotal
                subtotal_item = item['quantity'] * item['price']
                c.drawString(470, y, f"Rp {subtotal_item:,}")
                
                y -= 20
            
            # Line separator
            y -= 5
            c.line(50, y, width-50, y)
            
            # Summary
            y -= 25
            subtotal = transaction['total'] / 1.1
            tax = transaction['total'] - subtotal
            
            c.setFont("Helvetica", 10)
            c.drawString(360, y, "Subtotal:")
            c.drawString(470, y, f"Rp {subtotal:,.0f}")
            
            y -= 20
            c.drawString(360, y, "Pajak (10%):")
            c.drawString(470, y, f"Rp {tax:,.0f}")
            
            # Line separator
            y -= 10
            c.line(350, y, width-50, y)
            
            # Total
            y -= 25
            c.setFont("Helvetica-Bold", 12)
            c.drawString(360, y, "TOTAL:")
            c.drawString(470, y, f"Rp {transaction['total']:,.0f}")
            
            # Payment details (if cash)
            if cash_amount > 0:
                y -= 30
                c.setFont("Helvetica", 10)
                c.drawString(360, y, "Bayar:")
                c.drawString(470, y, f"Rp {cash_amount:,.0f}")
                
                y -= 20
                change = cash_amount - transaction['total']
                c.drawString(360, y, "Kembali:")
                c.drawString(470, y, f"Rp {change:,.0f}")
            
            # Footer
            y -= 50
            c.setFont("Helvetica-Bold", 11)
            c.drawCentredString(width/2, y, "Terima Kasih")
            
            y -= 20
            c.setFont("Helvetica", 9)
            c.drawCentredString(width/2, y, "Selamat Datang Kembali")
            
            # Save PDF
            c.save()
            
            # Show notification that PDF is saved
            self.show_notification(f"✓ PDF tersimpan: {filename}", duration=3000, type="success")
            
            # Ask to open PDF (PDF already saved regardless of answer)
            if messagebox.askyesno("PDF Tersimpan", f"Nota PDF berhasil disimpan!\n\nFile: {filename}\n\nApakah Anda ingin membuka PDF sekarang?"):
                os.startfile(filename)
            else:
                messagebox.showinfo("Info", f"PDF sudah tersimpan di:\n{os.path.abspath(filename)}\n\nAnda dapat membukanya kapan saja.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan nota: {str(e)}")
    
    def print_receipt_direct(self, transaction, cash_amount):
        """Print nota langsung ke printer (simplified - open file)"""
        try:
            filename = f"nota_{transaction['id']}.txt"
            
            # Save first
            self.save_receipt_pdf(transaction, cash_amount, None)
            
            # Open file with default program
            os.startfile(filename)
            
            self.show_notification("Nota dibuka untuk print", duration=2000)
            
        except Exception as e:
            messagebox.showerror("Error", f"Gagal membuka nota: {str(e)}")

    def show_pemasukan_page(self):
        self.btn_kasir.config(bg=self.colors['bg_card'], fg=self.colors['text_gray'])
        self.btn_pemasukan.config(bg="white", fg=self.colors['accent'])
        
        # Hide cart sidebar (only show in kasir page)
        if self.cart_sidebar:
            self.cart_sidebar.destroy()
            self.cart_sidebar = None
        
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        tk.Label(self.content_frame, text="Laporan Pemasukan", font=("Arial", 20, "bold"),
                bg=self.colors['bg_dark'], fg="white").pack(pady=20, padx=20, anchor="w")
        
        # Stats cards
        stats_frame = tk.Frame(self.content_frame, bg=self.colors['bg_dark'])
        stats_frame.pack(fill=tk.X, padx=20, pady=10)
        
        # Get stats from database or calculate from transactions
        if self.db:
            try:
                stats = get_transaction_stats(self.db)
                total_income = stats['total_income']
                total_trans = stats['total_transactions']
                avg_trans = stats['avg_transaction']
            except:
                total_income = sum(t["total"] for t in self.transactions)
                total_trans = len(self.transactions)
                avg_trans = total_income / total_trans if total_trans > 0 else 0
        else:
            total_income = sum(t["total"] for t in self.transactions)
            total_trans = len(self.transactions)
            avg_trans = total_income / total_trans if total_trans > 0 else 0
        
        self.create_stat_card(stats_frame, "💵", "Total Pemasukan", 
                             f"Rp {total_income:,.0f}", 0)
        self.create_stat_card(stats_frame, "🛒", "Total Transaksi", str(total_trans), 1)
        self.create_stat_card(stats_frame, "📈", "Rata-rata Transaksi", 
                             f"Rp {avg_trans:,.0f}", 2)
        
        # Table - Optimized height for 1366x768 screens
        table_frame = tk.Frame(self.content_frame, bg=self.colors['bg_card'], height=400)
        table_frame.pack(fill=tk.BOTH, padx=20, pady=10)
        table_frame.pack_propagate(False)
        
        tk.Label(table_frame, text="Riwayat Transaksi", font=("Arial", 14, "bold"),
                bg=self.colors['bg_card'], fg="white").pack(pady=15, padx=20, anchor="w")
        
        # Treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", 
                       background=self.colors['bg_dark'], 
                       foreground="white", 
                       fieldbackground=self.colors['bg_dark'],
                       borderwidth=0,
                       rowheight=32,
                       font=("Segoe UI", 9))
        style.configure("Treeview.Heading", 
                       background=self.colors['bg_card'],
                       foreground="white", 
                       borderwidth=0,
                       font=("Segoe UI", 10, "bold"),
                       relief="flat")
        style.map("Treeview", background=[("selected", self.colors['accent'])])
        
        tree = ttk.Treeview(table_frame, columns=("No", "Tanggal", "Jam", "Customer", "Item", "Metode", "Total"),
                           show="headings", height=15)
        
        tree.heading("No", text="No")
        tree.heading("Tanggal", text="Tanggal")
        tree.heading("Jam", text="Jam")
        tree.heading("Customer", text="Customer")
        tree.heading("Item", text="Item Pesanan")
        tree.heading("Metode", text="Pembayaran")
        tree.heading("Total", text="Total")
        
        tree.column("No", width=40, anchor="center")
        tree.column("Tanggal", width=95, anchor="center")
        tree.column("Jam", width=75, anchor="center")
        tree.column("Customer", width=110, anchor="center")
        tree.column("Item", width=290, anchor="center")
        tree.column("Metode", width=100, anchor="center")
        tree.column("Total", width=110, anchor="center")
        
        payment_names = {
            'cash': 'Tunai', 'debit': 'Debit', 'credit': 'Kredit',
            'ewallet': 'E-Wallet', 'qris': 'QRIS', 'transfer': 'Transfer'
        }
        
        for idx, trans in enumerate(reversed(self.transactions), 1):
            # Parse date and time
            date_obj = datetime.strptime(trans["date"], '%Y-%m-%d %H:%M:%S')
            date_str = date_obj.strftime('%d-%m-%Y')
            time_str = date_obj.strftime('%H:%M:%S')
            
            # Format items - lebih ringkas
            items_list = []
            for item in trans['items']:
                qty = item['quantity']
                name = item['name']
                # Singkat nama jika terlalu panjang
                if len(name) > 15:
                    name = name[:12] + "..."
                items_list.append(f"{name} ({qty}x)")
            items_text = ", ".join(items_list)
            
            # Truncate if too long
            if len(items_text) > 45:
                items_text = items_text[:42] + "..."
            
            customer_name = trans.get('customer_name', 'Umum')
            # Capitalize customer name
            if customer_name:
                customer_name = customer_name.title()
            
            tree.insert("", "end", values=(
                idx,
                date_str,
                time_str,
                customer_name,
                items_text,
                payment_names.get(trans["payment_method"], trans["payment_method"]),
                f"Rp {trans['total']:,.0f}"
            ))
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=10, padx=(0, 20))
        
        # Bind double-click event to show detail
        def on_double_click(event):
            selected_item = tree.selection()
            if selected_item:
                item_values = tree.item(selected_item[0])['values']
                if item_values:
                    # Get transaction index (No - 1 because we enumerate from 1)
                    trans_idx = len(self.transactions) - int(item_values[0])
                    if 0 <= trans_idx < len(self.transactions):
                        self.show_transaction_detail(self.transactions[trans_idx])
        
        tree.bind("<Double-Button-1>", on_double_click)
        
        # Bottom section (outside table_frame) - Compact for small screens
        bottom_section = tk.Frame(self.content_frame, bg=self.colors['bg_dark'])
        bottom_section.pack(fill=tk.X, padx=20, pady=10)
        
        # Left: Tips
        tips_frame = tk.Frame(bottom_section, bg=self.colors['bg_dark'])
        tips_frame.pack(side=tk.LEFT)
        
        tk.Label(tips_frame, text="💡", font=("Segoe UI", 12),
                bg=self.colors['bg_dark'], fg="#FFA500").pack(side=tk.LEFT, padx=(5, 5))
        
        tk.Label(tips_frame, text="Double-click baris untuk lihat detail", 
                font=("Segoe UI", 9),
                bg=self.colors['bg_dark'], fg=self.colors['text_gray']).pack(side=tk.LEFT)
        
        # Right: Export buttons (horizontal, compact)
        export_frame = tk.Frame(bottom_section, bg=self.colors['bg_dark'])
        export_frame.pack(side=tk.RIGHT)
        
        tk.Label(export_frame, text="📥 Export Laporan:", font=("Segoe UI", 9, "bold"),
                bg=self.colors['bg_dark'], fg="white").pack(side=tk.LEFT, padx=(0, 8))
        
        # Excel button
        tk.Button(export_frame, text="📊 Excel", font=("Segoe UI", 8, "bold"),
                 bg="#28a745", fg="white", relief=tk.FLAT,
                 command=self.export_to_excel, cursor="hand2",
                 padx=12, pady=6).pack(side=tk.LEFT, padx=2)
        
        # PDF button
        tk.Button(export_frame, text="📄 PDF", font=("Segoe UI", 8, "bold"),
                 bg=self.colors['accent'], fg="white", relief=tk.FLAT,
                 command=self.export_to_pdf, cursor="hand2",
                 padx=12, pady=6).pack(side=tk.LEFT, padx=2)
    
    def show_transaction_detail(self, transaction):
        """Tampilkan detail lengkap transaksi"""
        # Create detail window
        detail_window = tk.Toplevel(self.root)
        detail_window.title(f"Detail Transaksi - {transaction['id']}")
        detail_window.geometry("600x700")
        detail_window.configure(bg=self.colors['bg_card'])
        detail_window.transient(self.root)
        
        # Center window
        detail_window.update_idletasks()
        x = (detail_window.winfo_screenwidth() // 2) - (detail_window.winfo_width() // 2)
        y = (detail_window.winfo_screenheight() // 2) - (detail_window.winfo_height() // 2)
        detail_window.geometry(f"+{x}+{y}")
        
        # Header
        header = tk.Frame(detail_window, bg=self.colors['bg_dark'])
        header.pack(fill=tk.X, padx=0, pady=0)
        
        header_content = tk.Frame(header, bg=self.colors['bg_dark'])
        header_content.pack(fill=tk.X, padx=20, pady=20)
        
        tk.Label(header_content, text="📋 Detail Transaksi", font=("Segoe UI", 16, "bold"),
                bg=self.colors['bg_dark'], fg="white").pack(side=tk.LEFT)
        
        tk.Button(header_content, text="×", font=("Arial", 20), bg=self.colors['bg_dark'],
                 fg="white", relief=tk.FLAT, command=detail_window.destroy,
                 cursor="hand2").pack(side=tk.RIGHT)
        
        # Content frame with scroll
        canvas = tk.Canvas(detail_window, bg=self.colors['bg_card'], highlightthickness=0)
        scrollbar = tk.Scrollbar(detail_window, orient="vertical", command=canvas.yview)
        content_frame = tk.Frame(canvas, bg=self.colors['bg_card'])
        
        content_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=content_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Enable mouse wheel scrolling
        self.bind_mousewheel(canvas)
        
        # Transaction info section
        info_section = tk.Frame(content_frame, bg=self.colors['bg_dark'])
        info_section.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(info_section, text="Informasi Transaksi", font=("Segoe UI", 12, "bold"),
                bg=self.colors['bg_dark'], fg="white", anchor="w").pack(fill=tk.X, pady=(10, 15), padx=15)
        
        # Transaction details
        details = [
            ("ID Transaksi", transaction['id']),
            ("Tanggal & Waktu", transaction['date']),
            ("Nama Customer", transaction.get('customer_name', 'Umum').title()),
            ("Metode Pembayaran", {
                'cash': 'Tunai', 'debit': 'Kartu Debit', 'credit': 'Kartu Kredit',
                'ewallet': 'E-Wallet', 'qris': 'QRIS', 'transfer': 'Transfer Bank'
            }.get(transaction['payment_method'], transaction['payment_method']))
        ]
        
        for label, value in details:
            detail_row = tk.Frame(info_section, bg=self.colors['bg_dark'])
            detail_row.pack(fill=tk.X, padx=15, pady=5)
            
            tk.Label(detail_row, text=label + ":", font=("Segoe UI", 10),
                    bg=self.colors['bg_dark'], fg=self.colors['text_gray'],
                    anchor="w", width=20).pack(side=tk.LEFT)
            
            tk.Label(detail_row, text=value, font=("Segoe UI", 10, "bold"),
                    bg=self.colors['bg_dark'], fg="white", anchor="w").pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Items section
        items_section = tk.Frame(content_frame, bg=self.colors['bg_dark'])
        items_section.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(items_section, text="Item Pesanan", font=("Segoe UI", 12, "bold"),
                bg=self.colors['bg_dark'], fg="white", anchor="w").pack(fill=tk.X, pady=(10, 15), padx=15)
        
        # Items table header
        header_row = tk.Frame(items_section, bg=self.colors['bg_card'])
        header_row.pack(fill=tk.X, padx=15, pady=(0, 5))
        
        tk.Label(header_row, text="Item", font=("Segoe UI", 9, "bold"),
                bg=self.colors['bg_card'], fg=self.colors['text_gray'],
                anchor="w", width=30).pack(side=tk.LEFT, padx=5)
        tk.Label(header_row, text="Qty", font=("Segoe UI", 9, "bold"),
                bg=self.colors['bg_card'], fg=self.colors['text_gray'],
                anchor="center", width=8).pack(side=tk.LEFT, padx=5)
        tk.Label(header_row, text="Harga", font=("Segoe UI", 9, "bold"),
                bg=self.colors['bg_card'], fg=self.colors['text_gray'],
                anchor="e", width=12).pack(side=tk.LEFT, padx=5)
        tk.Label(header_row, text="Subtotal", font=("Segoe UI", 9, "bold"),
                bg=self.colors['bg_card'], fg=self.colors['text_gray'],
                anchor="e", width=12).pack(side=tk.LEFT, padx=5)
        
        # Items list
        for item in transaction['items']:
            item_row = tk.Frame(items_section, bg="#1a1a1a")
            item_row.pack(fill=tk.X, padx=15, pady=3)
            
            tk.Label(item_row, text=item['name'], font=("Segoe UI", 10),
                    bg="#1a1a1a", fg="white", anchor="w", width=30).pack(side=tk.LEFT, padx=5, pady=8)
            tk.Label(item_row, text=str(item['quantity']), font=("Segoe UI", 10),
                    bg="#1a1a1a", fg="white", anchor="center", width=8).pack(side=tk.LEFT, padx=5)
            tk.Label(item_row, text=f"Rp {item['price']:,}", font=("Segoe UI", 10),
                    bg="#1a1a1a", fg="white", anchor="e", width=12).pack(side=tk.LEFT, padx=5)
            tk.Label(item_row, text=f"Rp {item['price'] * item['quantity']:,}", font=("Segoe UI", 10, "bold"),
                    bg="#1a1a1a", fg="white", anchor="e", width=12).pack(side=tk.LEFT, padx=5)
        
        # Summary section
        summary_section = tk.Frame(content_frame, bg=self.colors['bg_dark'])
        summary_section.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(summary_section, text="Ringkasan Pembayaran", font=("Segoe UI", 12, "bold"),
                bg=self.colors['bg_dark'], fg="white", anchor="w").pack(fill=tk.X, pady=(10, 15), padx=15)
        
        subtotal = transaction['total'] / 1.1
        tax = transaction['total'] - subtotal
        
        summary_items = [
            ("Subtotal", f"Rp {subtotal:,.0f}", "white"),
            ("Pajak (10%)", f"Rp {tax:,.0f}", "white"),
            ("TOTAL", f"Rp {transaction['total']:,.0f}", self.colors['accent'])
        ]
        
        for label, value, color in summary_items:
            summary_row = tk.Frame(summary_section, bg=self.colors['bg_dark'])
            summary_row.pack(fill=tk.X, padx=15, pady=5)
            
            font_style = ("Segoe UI", 14, "bold") if label == "TOTAL" else ("Segoe UI", 10)
            
            tk.Label(summary_row, text=label + ":", font=font_style,
                    bg=self.colors['bg_dark'], fg=self.colors['text_gray'] if label != "TOTAL" else "white",
                    anchor="w").pack(side=tk.LEFT)
            
            tk.Label(summary_row, text=value, font=font_style,
                    bg=self.colors['bg_dark'], fg=color, anchor="e").pack(side=tk.RIGHT)
        
        # Action buttons
        button_frame = tk.Frame(content_frame, bg=self.colors['bg_card'])
        button_frame.pack(fill=tk.X, padx=20, pady=20)
        
        tk.Button(button_frame, text="🖨️ Cetak Ulang Nota", font=("Segoe UI", 11, "bold"),
                 bg=self.colors['accent'], fg="white", relief=tk.FLAT,
                 command=lambda: self.print_receipt(transaction, 0),
                 cursor="hand2", pady=12).pack(fill=tk.X, pady=5)
        
        tk.Button(button_frame, text="Tutup", font=("Segoe UI", 11),
                 bg=self.colors['bg_dark'], fg="white", relief=tk.FLAT,
                 command=detail_window.destroy, cursor="hand2", pady=12).pack(fill=tk.X, pady=5)
    
    def export_to_excel(self):
        """Export laporan transaksi ke Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Laporan Transaksi"
            
            # Header styling
            header_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:H1')
            ws['A1'] = "LAPORAN TRANSAKSI - KEDAI HAUNA"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Subtitle
            ws.merge_cells('A2:H2')
            ws['A2'] = f"Tanggal Export: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Headers
            headers = ["No", "ID Transaksi", "Tanggal", "Jam", "Customer", "Item", "Metode", "Total"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Data
            payment_names = {
                'cash': 'Tunai', 'debit': 'Debit', 'credit': 'Kredit',
                'ewallet': 'E-Wallet', 'qris': 'QRIS', 'transfer': 'Transfer'
            }
            
            row = 5
            for idx, trans in enumerate(reversed(self.transactions), 1):
                date_obj = datetime.strptime(trans["date"], '%Y-%m-%d %H:%M:%S')
                date_str = date_obj.strftime('%d-%m-%Y')
                time_str = date_obj.strftime('%H:%M:%S')
                
                items_list = [f"{item['name']} ({item['quantity']}x)" for item in trans['items']]
                items_text = ", ".join(items_list)
                
                customer_name = str(trans.get('customer_name', 'Umum')).title()
                
                # Convert Decimal to float for Excel compatibility
                total_value = float(trans['total']) if hasattr(trans['total'], '__float__') else trans['total']
                
                data = [
                    idx,
                    str(trans['id']),
                    date_str,
                    time_str,
                    customer_name,
                    items_text,
                    payment_names.get(trans['payment_method'], trans['payment_method']),
                    total_value
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col == 8:  # Total column
                        cell.number_format = 'Rp #,##0'
                        cell.alignment = Alignment(horizontal='right')
                    elif col in [1, 3, 4, 7]:  # Center align
                        cell.alignment = Alignment(horizontal='center')
                
                row += 1
            
            # Summary
            row += 1
            # Convert to float for sum calculation
            total_income = sum(float(t["total"]) if hasattr(t["total"], '__float__') else t["total"] for t in self.transactions)
            total_trans = len(self.transactions)
            
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "TOTAL PEMASUKAN"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].alignment = Alignment(horizontal='right')
            ws[f'H{row}'] = total_income
            ws[f'H{row}'].font = Font(bold=True, size=12)
            ws[f'H{row}'].number_format = 'Rp #,##0'
            ws[f'H{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            row += 1
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "TOTAL TRANSAKSI"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].alignment = Alignment(horizontal='right')
            ws[f'H{row}'] = total_trans
            ws[f'H{row}'].font = Font(bold=True)
            ws[f'H{row}'].alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 40
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            
            # Save file
            filename = f"laporan_transaksi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            self.show_notification(f"✓ Excel tersimpan: {filename}", duration=3000, type="success")
            
            if messagebox.askyesno("Excel Tersimpan", f"Laporan Excel berhasil disimpan!\n\nFile: {filename}\n\nApakah Anda ingin membuka file sekarang?"):
                os.startfile(filename)
            else:
                messagebox.showinfo("Info", f"File Excel sudah tersimpan di:\n{os.path.abspath(filename)}")
                
        except ImportError:
            messagebox.showerror("Error", "Library openpyxl belum terinstall!\n\nJalankan: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal export ke Excel: {str(e)}")
    
    def export_to_pdf(self):
        """Export laporan transaksi ke PDF"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import mm
            from reportlab.pdfgen import canvas
            from reportlab.lib import colors
            from reportlab.platypus import Table, TableStyle
            
            filename = f"laporan_transaksi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Create PDF in landscape mode for better table view
            c = canvas.Canvas(filename, pagesize=landscape(A4))
            width, height = landscape(A4)
            
            # Header
            y = height - 40
            c.setFont("Helvetica-Bold", 18)
            c.drawCentredString(width/2, y, "LAPORAN TRANSAKSI - KEDAI HAUNA")
            
            y -= 20
            c.setFont("Helvetica", 10)
            c.drawCentredString(width/2, y, f"Tanggal Export: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
            
            # Line
            y -= 15
            c.line(40, y, width-40, y)
            
            # Summary stats
            y -= 30
            total_income = sum(t["total"] for t in self.transactions)
            total_trans = len(self.transactions)
            avg_trans = total_income / total_trans if total_trans > 0 else 0
            
            c.setFont("Helvetica-Bold", 11)
            c.drawString(40, y, f"Total Pemasukan: Rp {total_income:,.0f}")
            c.drawString(250, y, f"Total Transaksi: {total_trans}")
            c.drawString(450, y, f"Rata-rata: Rp {avg_trans:,.0f}")
            
            # Table header
            y -= 30
            c.setFont("Helvetica-Bold", 9)
            headers = ["No", "ID Transaksi", "Tanggal", "Jam", "Customer", "Item", "Metode", "Total"]
            x_positions = [40, 70, 160, 220, 270, 340, 550, 620]
            
            for i, header in enumerate(headers):
                c.drawString(x_positions[i], y, header)
            
            y -= 5
            c.line(40, y, width-40, y)
            
            # Data
            y -= 20
            c.setFont("Helvetica", 8)
            
            payment_names = {
                'cash': 'Tunai', 'debit': 'Debit', 'credit': 'Kredit',
                'ewallet': 'E-Wallet', 'qris': 'QRIS', 'transfer': 'Transfer'
            }
            
            for idx, trans in enumerate(reversed(self.transactions), 1):
                if y < 60:  # New page if needed
                    c.showPage()
                    y = height - 40
                    c.setFont("Helvetica", 8)
                
                date_obj = datetime.strptime(trans["date"], '%Y-%m-%d %H:%M:%S')
                date_str = date_obj.strftime('%d-%m-%Y')
                time_str = date_obj.strftime('%H:%M')
                
                items_list = [f"{item['name']} ({item['quantity']}x)" for item in trans['items']]
                items_text = ", ".join(items_list)
                if len(items_text) > 35:
                    items_text = items_text[:32] + "..."
                
                customer_name = trans.get('customer_name', 'Umum').title()
                
                c.drawString(x_positions[0], y, str(idx))
                c.drawString(x_positions[1], y, trans['id'][:12])
                c.drawString(x_positions[2], y, date_str)
                c.drawString(x_positions[3], y, time_str)
                c.drawString(x_positions[4], y, customer_name[:12])
                c.drawString(x_positions[5], y, items_text)
                c.drawString(x_positions[6], y, payment_names.get(trans['payment_method'], '')[:10])
                c.drawString(x_positions[7], y, f"Rp {trans['total']:,.0f}")
                
                y -= 18
            
            # Footer line
            y -= 10
            c.line(40, y, width-40, y)
            
            # Save
            c.save()
            
            self.show_notification(f"✓ PDF tersimpan: {filename}", duration=3000, type="success")
            
            if messagebox.askyesno("PDF Tersimpan", f"Laporan PDF berhasil disimpan!\n\nFile: {filename}\n\nApakah Anda ingin membuka file sekarang?"):
                os.startfile(filename)
            else:
                messagebox.showinfo("Info", f"File PDF sudah tersimpan di:\n{os.path.abspath(filename)}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Gagal export ke PDF: {str(e)}")
    
    def create_stat_card(self, parent, icon, title, value, col):
        card = tk.Frame(parent, bg=self.colors['bg_card'])
        card.grid(row=0, column=col, padx=8, pady=8, sticky="nsew")
        parent.grid_columnconfigure(col, weight=1)
        
        tk.Label(card, text=icon, font=("Arial", 28), bg=self.colors['bg_dark'],
                fg="white").pack(pady=10)
        tk.Label(card, text=title, font=("Arial", 9), bg=self.colors['bg_card'],
                fg=self.colors['text_gray']).pack()
        tk.Label(card, text=value, font=("Arial", 16, "bold"), bg=self.colors['bg_card'],
                fg="white").pack(pady=8)
    
    def load_transactions(self):
        """Load transactions from database or JSON fallback"""
        if self.db:
            try:
                return get_all_transactions(self.db)
            except Exception as e:
                print(f"Error loading from database: {e}")
                messagebox.showwarning("Database Warning", 
                                     "Gagal load dari database, menggunakan JSON backup")
        
        # Fallback to JSON
        if os.path.exists("transactions.json"):
            with open("transactions.json", "r") as f:
                return json.load(f)
        return []
    
    def save_transactions(self):
        """Save transactions to JSON as backup"""
        try:
            # Convert Decimal to float for JSON serialization
            def convert_decimal(obj):
                if hasattr(obj, '__float__'):
                    return float(obj)
                elif isinstance(obj, dict):
                    return {k: convert_decimal(v) for k, v in obj.items()}
                elif isinstance(obj, list):
                    return [convert_decimal(item) for item in obj]
                return obj
            
            transactions_json = convert_decimal(self.transactions)
            
            with open("transactions.json", "w") as f:
                json.dump(transactions_json, f, indent=2)
        except Exception as e:
            print(f"Error saving to JSON: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = KedaiHaunaApp(root)
    root.mainloop()
